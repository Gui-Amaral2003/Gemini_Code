from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from typing import Optional

SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}

def _validate_spreadsheet_path(path: str) -> tuple[Optional[Path], Optional[str]]:
    """Validações comuns antes de tentar abrir a planilha. Retorna (Path, error_message)"""
    file_path = Path(path)

    if not file_path.exists():
        return None, f"Arquivo não encontrado: {path}"

    if not file_path.is_file():
        return None, f"O caminho não é um arquivo: {path}"

    if file_path.suffix.lower() not in SPREADSHEET_EXTENSIONS:
        return None, (
            f"Extensão não suportada: '{file_path.suffix}'. "
            f"Extensões aceitas: {', '.join(sorted(SPREADSHEET_EXTENSIONS))}")

    return file_path, None

def _is_csv(file_path: Path) -> bool:
    return file_path.suffix.lower() == ".csv"

def _load_dataframe(file_path: Path, sheet_name: Optional[str]) -> tuple[Optional[pd.DataFrame], Optional[str]]:
    """Carrega a tabela como DataFrame. Para CSV, sheet_name é ignorado. Para Excel, se sheet_name não for informado usa a primeira aba. Retorna (dataframe, error_message)"""
    try:
        if _is_csv(file_path):
            return pd.read_csv(file_path), None

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = workbook.sheetnames
        workbook.close()

        if not available_sheets:
            return None, f"Arquivo Excel não contém abas: {file_path}"

        target_sheet = sheet_name or available_sheets[0]

        if target_sheet not in available_sheets:
            return None, (
                f"Aba '{target_sheet}' não encontrada no arquivo Excel. "
                f"Abas disponíveis: {', '.join(available_sheets)}"
            )

        return pd.read_excel(file_path, sheet_name=target_sheet), None

    except Exception as e:
        return None, f"Erro ao carregar a planilha: {e}"

def _validate_columns(df: pd.DataFrame, columns: Optional[list[str]]) -> tuple[Optional[list[str]], Optional[str]]:
    """Confere se as colunas pedidas existem. Erro já traz as colunas válidas. Retorna (columns_to_use, error_message)"""

    if not columns:
        return None, None  # Se não pediu colunas, não precisa validar

    invalid = [c for c in columns if c not in df.columns]

    if invalid:
        return None, (
            f"Colunas inválidas: {', '.join(invalid)}. "
            f"Colunas disponíveis: {', '.join(df.columns)}"
        )

    return columns, None

def list_sheets(path: str) -> str:
    """Lista as abas de uma planilha Excel com suas dimensões (linhas x colunas). Para CSV, retorna apenas as dimensões da tabela. Util para modelo e usuario se orientar antes de ler ou buscar dados"""

    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    try:
        if _is_csv(file_path):
            df = pd.read_csv(file_path)
            return (
                f"Arquivo CSV: {file_path.name}\n"
                f"Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas\n"
                f"Colunas: {', '.join(df.columns)}"
            )

        workbook = load_workbook(file_path, read_only=True, data_only=True)

        if not workbook.sheetnames:
            return f"Arquivo Excel não contém abas: {file_path}"

        lines = [
            f"- {name}: {workbook[name].max_row} linha(s) x "
            f"{workbook[name].max_column} coluna(s)"
            for name in workbook.sheetnames
        ]
        workbook.close()

        return 'Abas disponíveis:\n' + '\n'.join(lines)

    except Exception as e:
        return f"Erro ao ler o arquivo: {path}. Detalhes: {e}"

def preview_sheet(path: str, sheet_name: Optional[str] = None, n_rows: int = 15) -> str:
    """Mostra as primeiras linhas de uma planilha Excel (ou CSV), além do total de linhas. Server para entender se vale ler tudo ou buscar algo especifico"""
    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    preview = df.head(n_rows)

    header = (
        f"Colunas: {', '.join(preview.columns)}\n"
        f"Total de linhas: {df.shape[0]}\n"
    )

    return header + preview.to_markdown(index = False)

def read_sheet(path: str, sheet_name: Optional[str] = None, start_row: int = 0, max_rows: int = 100, columns: Optional[list[str]] = None) -> str:
    """Lê os dados de uma aba (ou CSV) de forma paginada, começando em start_row e trazendo no máximo max_rows linhas. Se columns for informado, retorna apenas essas colunas. Indica no final se há mais linhas disponíveis e qual start_row usar para continuar a leitura"""

    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    selected_columns, error = _validate_columns(df, columns)
    if error:
        return error

    if selected_columns:
        df = df[selected_columns]

    total_rows = len(df)

    if start_row >= total_rows:
        return (
            f"start_row ({start_row}) é maior ao total de linhas ({total_rows}). "
            f"Não há mais dados para ler."
        )

    end_row = min(start_row + max_rows, total_rows)
    page = df.iloc[start_row:end_row]

    footer = f"\n\nMostrando linhas {start_row} a {end_row - 1} de {total_rows}."

    if end_row < total_rows:
        footer += (
            f" Para continuar a leitura, use start_row={end_row} na próxima chamada."
        )

    return page.to_markdown(index = False) + footer

def search_in_sheet(path: str, query: str, sheet_name: Optional[str] = None, column: Optional[list[str]] = None, max_matches: int = 30) -> str:
    """Busca um valor (texto ou número, comparação por substring, sem diferenciar maiúsculas/minúsculas) em uma aba/CSV. Se column for informado, busca só nessa coluna; caso contrário, busca em todas as colunas. Retorna as linhas que batem, limitado a max_matches."""
    
    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    if column:
        _, error = _validate_columns(df, column)
        if error:
            return error
        target = df[[column]]
    else:
        target = df

    mask = target.apply(
        lambda col: col.astype(str).str.contains(query, case = False, na = False)
    ).any(axis = 1)

    matches = df[mask]
    total_matches = len(matches)

    if total_matches == 0:
        return f"Nenhuma correspondência encontrada para '{query}'."

    shown = matches.head(max_matches)

    footer = f"\n\n{total_matches} linha(s) encontrada(s)."
    if total_matches > max_matches:
        footer += f" Mostrando as primeiras {max_matches}."
 
    return shown.to_markdown(index=False) + footer

