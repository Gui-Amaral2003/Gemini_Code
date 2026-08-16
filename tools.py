from pathlib import Path
import re
import os
import pandas as pd
from openpyxl import load_workbook
from typing import Optional
# --------------------------------------------------------------------------- #
# Ler arquivos do próprio sistema
# --------------------------------------------------------------------------- #

def read_file(path: str) -> str:
    """Lê um arquivo de texto."""

    file_path = Path(path)

    if not file_path.exists():
        return f"Arquivo não encontrado: {path}"

    if not file_path.is_file():
        return f"O caminho não é um arquivo: {path}"

    try:
        return file_path.read_text(encoding="utf-8")

    except UnicodeDecodeError:
        return f"Erro ao decodificar o arquivo: {path}. Certifique-se de que está em UTF-8."

    except OSError as e:
        return f"Erro ao ler o arquivo: {path}. Detalhes: {e}"

# --------------------------------------------------------------------------- #
# Funções para acesso ao banco de dados
# --------------------------------------------------------------------------- #
TABELAS_PERMITIDAS = {
    "pedidos": {
        "schema": "vendas",
        "colunas_filtro": {"cliente_id": "int", "status": "str", "data_pedido": "date"},
        "colunas_retorno": ["id", "cliente_id", "status", "data_pedido", "valor"],
    },
    "clientes": {
        "schema": "vendas",
        "colunas_filtro": {"cliente_id": "int", "cidade": "str"},
        "colunas_retorno": ["id", "nome", "cidade"],  # note: sem CPF, mesmo que exista na tabela
    },
}

_OPERADORES_POR_TIPO = {
    "int": {"=", "!=", ">", "<", ">=", "<="},
    "float": {"=", "!=", ">", "<", ">=", "<="},
    "date": {"=", "!=", ">", "<", ">=", "<="},
    "str": {"=", "!=", "LIKE"}
}

# Validação extra de indentificadores válidos
_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

MAX_ROWS = 10_000
QUERY_TIMEOUT_SECONDS = 10

class QueryValidationError(ValueError):
    """Erro de validação de query."""

def _validate_identifier(name: str, label: str) -> None:
    if not _IDENTIFIER_RE.match(name):
        raise QueryValidationError(f"{label} inválido: {name!r}")

def _coerce_value(value, tipo: str, coluna: str):
    """Converte o valor para o tipo esperado"""
    try:
        if tipo == 'int':
            return int(value)
        if tipo == 'float':
            return float(value)
        if tipo == 'date':
            # Aceita YYYY-MM-DD, deixa o driver validar o formato final
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", str(value)):
                raise ValueError
            return str(value)
        if tipo == 'str':
            return str(value)

    except (ValueError, TypeError) as e:
        raise QueryValidationError(f"Erro ao converter valor para coluna {coluna}: {value}. Detalhes: {e}")

def _build_query(table: str, conditions: list[dict]):
    """
    Valida table + conditions contra o catálogo e monta uma query parametrizada.

    Retorna (sql, params)
    """
    if table not in TABELAS_PERMITIDAS:
        raise QueryValidationError(
            f"Tabela não permitida: {table}"
            f"Disponíveis: {list(TABELAS_PERMITIDAS)}"
        )

    config = TABELAS_PERMITIDAS[table]
    schema = config["schema"]
    colunas_filtro = config["colunas_filtro"]
    colunas_retorno = config['colunas_retorno']

    _validate_identifier(schema, 'schema')
    _validate_identifier(table, 'tabela')
    for col in colunas_retorno:
        _validate_identifier(col, 'coluna de retorno')

    where_clauses = []
    params = []

    for cond in conditions or []:
        coluna = cond.get('column')
        operador = cond.get('operator')
        valor = cond.get('value')

        if coluna not in colunas_filtro:
            raise QueryValidationError(
                f"Coluna não permitida para filtro: {coluna}. "
                f"Disponíveis: {list(colunas_filtro)}"
            )

        _validate_identifier(coluna, 'coluna de filtro')

        tipo = colunas_filtro[coluna]
        operadores_validos = _OPERADORES_POR_TIPO[tipo]

        if operador not in operadores_validos:
            raise QueryValidationError(
                f"Operador '{operador}' não é permitido para a coluna '{coluna}' "
                f"(tipo {tipo}). Permitidos: {sorted(operadores_validos)}"
            )

        valor_validado = _coerce_value(valor, tipo, coluna)

        if operador == "LIKE":
            # o próprio valor pode ter % embutido; se não tiver, envolve automaticamente
            if "%" not in valor_validado:
                valor_validado = f"%{valor_validado}%"

        where_clauses.append(f"[{coluna}] {operador} :param_{len(params)}")
        params.append(valor_validado)

    colunas_sql = ', '.join(f"[{c}]" for c in colunas_retorno)
    sql = f'SELECT TOP {MAX_ROWS} {colunas_sql} FROM [{schema}].[{table}]'

    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)

    return sql, params

def query_table(table: str, conditions: list[dict] | None = None) -> str:
    """
    Executa uma consulta SELECT em uma tabela pré-cadastrada, filtrando apenas por colunas e operadores permitidos no catálogo TABELAS_PERMITIDAS
    """

    try:
        sql, params = _build_query(table, conditions or [])
    except QueryValidationError as e:
        return f"Requisição rejeitada: {e}"

    try:
        import sqlalchemy
        conn_string = os.environ.get("DB_CONN_STRING")

        engine = sqlalchemy.create_engine(conn_string, connect_args={"timeout": QUERY_TIMEOUT_SECONDS})

        with engine.connect() as conn:

            params_dict = {
                f"param_{i}": value
                for i, value in enumerate(params)
            }

            result = conn.execute(sqlalchemy.text(sql), params_dict)
            rows = result.fetchall()

            if not rows:
                return "Nenhum resultado encontrado."

            # Converte para lista de dicionários
            result_list = [dict(row._mapping) for row in rows]
            return str(result_list)
    except Exception as e:
        return f"Erro ao executar a query: {e}"

# -------------------------------------------------------------------------- #
# Funções para leitura de arquivos Excel e CSV
# -------------------------------------------------------------------------- #
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}

def _validate_spreadsheet_path(path: str) -> tuple[Optional[Path], Optional[str]]:
    """Validações comuns antes de tentar abrir a planilha. Retorna (Path, error_message)"""
    file_path = Path(path)

    if not file_path.exists():
        return None, f"Arquivo não encontrado: {path}"

    if not file_path.is_file():
        return None, f"O caminho não é um arquivo: {path}"

    if file_path.suffix.lower() not in SPREADSHEET_EXTENSIONS:
        return None, (
            f"Extensão não suportada: '{file_path.suffix}'. "
            f"Extensões aceitas: {', '.join(sorted(SPREADSHEET_EXTENSIONS))}")

    return file_path, None

def _is_csv(file_path: Path) -> bool:
    return file_path.suffix.lower() == ".csv"

def _load_dataframe(file_path: Path, sheet_name: Optional[str]) -> tuple[Optional[pd.DataFrame], Optional[str]]:
    """Carrega a tabela como DataFrame. Para CSV, sheet_name é ignorado. Para Excel, se sheet_name não for informado usa a primeira aba. Retorna (dataframe, error_message)"""
    try:
        if _is_csv(file_path):
            return pd.read_csv(file_path), None

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = workbook.sheetnames
        workbook.close()

        if not available_sheets:
            return None, f"Arquivo Excel não contém abas: {file_path}"

        target_sheet = sheet_name or available_sheets[0]

        if target_sheet not in available_sheets:
            return None, (
                f"Aba '{target_sheet}' não encontrada no arquivo Excel. "
                f"Abas disponíveis: {', '.join(available_sheets)}"
            )

        return pd.read_excel(file_path, sheet_name=target_sheet), None

    except Exception as e:
        return None, f"Erro ao carregar a planilha: {e}"

def _validate_columns(df: pd.DataFrame, columns: Optional[list[str]]) -> tuple[Optional[list[str]], Optional[str]]:
    """Confere se as colunas pedidas existem. Erro já traz as colunas válidas. Retorna (columns_to_use, error_message)"""

    if not columns:
        return None, None  # Se não pediu colunas, não precisa validar

    invalid = [c for c in columns if c not in df.columns]

    if invalid:
        return None, (
            f"Colunas inválidas: {', '.join(invalid)}. "
            f"Colunas disponíveis: {', '.join(df.columns)}"
        )

    return columns, None

def list_sheets(path: str) -> str:
    """Lista as abas de uma planilha Excel com suas dimensões (linhas x colunas). Para CSV, retorna apenas as dimensões da tabela. Util para modelo e usuario se orientar antes de ler ou buscar dados"""

    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    try:
        if _is_csv(file_path):
            df = pd.read_csv(file_path)
            return (
                f"Arquivo CSV: {file_path.name}\n"
                f"Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas\n"
                f"Colunas: {', '.join(df.columns)}"
            )

        workbook = load_workbook(file_path, read_only=True, data_only=True)

        if not workbook.sheetnames:
            return f"Arquivo Excel não contém abas: {file_path}"

        lines = [
            f"- {name}: {workbook[name].max_row} linha(s) x "
            f"{workbook[name].max_column} coluna(s)"
            for name in workbook.sheetnames
        ]
        workbook.close()

        return 'Abas disponíveis:\n' + '\n'.join(lines)

    except Exception as e:
        return f"Erro ao ler o arquivo: {path}. Detalhes: {e}"

def preview_sheet(path: str, sheet_name: Optional[str] = None, n_rows: int = 15) -> str:
    """Mostra as primeiras linhas de uma planilha Excel (ou CSV), além do total de linhas. Server para entender se vale ler tudo ou buscar algo especifico"""
    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    preview = df.head(n_rows)

    header = (
        f"Colunas: {', '.join(preview.columns)}\n"
        f"Total de linhas: {df.shape[0]}\n"
    )

    return header + preview.to_markdown(index = False)

def read_sheet(path: str, sheet_name: Optional[str] = None, start_row: int = 0, max_rows: int = 100, columns: Optional[list[str]] = None) -> str:
    """Lê os dados de uma aba (ou CSV) de forma paginada, começando em start_row e trazendo no máximo max_rows linhas. Se columns for informado, retorna apenas essas colunas. Indica no final se há mais linhas disponíveis e qual start_row usar para continuar a leitura"""

    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    selected_columns, error = _validate_columns(df, columns)
    if error:
        return error

    if selected_columns:
        df = df[selected_columns]

    total_rows = len(df)

    if start_row >= total_rows:
        return (
            f"start_row ({start_row}) é maior ao total de linhas ({total_rows}). "
            f"Não há mais dados para ler."
        )

    end_row = min(start_row + max_rows, total_rows)
    page = df.iloc[start_row:end_row]

    footer = f"\n\nMostrando linhas {start_row} a {end_row - 1} de {total_rows}."

    if end_row < total_rows:
        footer += (
            f" Para continuar a leitura, use start_row={end_row} na próxima chamada."
        )

    return page.to_markdown(index = False) + footer

def search_in_sheet(path: str, query: str, sheet_name: Optional[str] = None, column: Optional[list[str]] = None, max_matches: int = 30) -> str:
    """Busca um valor (texto ou número, comparação por substring, sem diferenciar maiúsculas/minúsculas) em uma aba/CSV. Se column for informado, busca só nessa coluna; caso contrário, busca em todas as colunas. Retorna as linhas que batem, limitado a max_matches."""
    
    file_path, error = _validate_spreadsheet_path(path)
    if error:
        return error

    df, error = _load_dataframe(file_path, sheet_name)
    if error:
        return error

    if df.empty:
        return f"A planilha/aba está vazia: {file_path.name}"

    if column:
        _, error = _validate_columns(df, column)
        if error:
            return error
        target = df[[column]]
    else:
        target = df

    mask = target.apply(
        lambda col: col.astype(str).str.contains(query, case = False, na = False)
    ).any(axis = 1)

    matches = df[mask]
    total_matches = len(matches)

    if total_matches == 0:
        return f"Nenhuma correspondência encontrada para '{query}'."

    shown = matches.head(max_matches)

    footer = f"\n\n{total_matches} linha(s) encontrada(s)."
    if total_matches > max_matches:
        footer += f" Mostrando as primeiras {max_matches}."
 
    return shown.to_markdown(index=False) + footer


# Funções que o Python realmente pode executar.
TOOLS = {
    "read_file": read_file,
    'query_table': query_table,
    'list_sheets': list_sheets,
    'preview_sheet': preview_sheet,
    'read_sheet': read_sheet,
    'search_in_sheet': search_in_sheet
}


# Descrição das ferramentas que será enviada ao Gemini.
TOOL_DEFINITIONS = [
    {
        "type": "function",
        "name": "read_file",
        "description": "Lê o conteúdo de um arquivo de texto.",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Caminho do arquivo que deve ser lido.",
                }
            },
            "required": ["path"],
        },
    },
    {
        "type": "function",
        "name": "query_table",
        "description": (
            "Consulta dados em uma tabela pré-cadastrada do banco de dados. "
            "Somente tabelas e colunas explicitamente permitidas podem ser "
            "usadas. Os filtros são combinados sempre com AND."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "table": {
                    "type": "string",
                    "description": "Nome da tabela a consultar.",
                    "enum": list(TABELAS_PERMITIDAS.keys()),
                },
                "conditions": {
                    "type": "array",
                    "description": "Lista de condições de filtro, combinadas com AND.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "column": {
                                "type": "string",
                                "description": "Nome da coluna a filtrar.",
                            },
                            "operator": {
                                "type": "string",
                                "description": "Operador de comparação.",
                                "enum": ["=", "!=", ">", "<", ">=", "<=", "LIKE"],
                            },
                            "value": {
                                "description": "Valor a comparar.",
                            },
                        },
                        "required": ["column", "operator", "value"],
                    },
                },
            },
            "required": ["table"],
        },
    },
    {
        "type": "function",
        "name": "list_sheets",
        "description": (
            "Lista as abas de um arquivo Excel (.xlsx/.xls) com suas dimensões, "
            "ou as dimensões da tabela caso seja um CSV. Use antes de ler ou "
            "buscar dados para descobrir os nomes das abas disponíveis."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Caminho do arquivo (.xlsx, .xls ou .csv).",
                }
            },
            "required": ["path"],
        },
    },
    {
        "type": "function",
        "name": "preview_sheet",
        "description": (
            "Mostra as colunas, o total de linhas e uma amostra das primeiras "
            "linhas de uma aba (ou de um CSV). Use para entender a estrutura "
            "da planilha antes de ler tudo ou fazer uma busca."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Caminho do arquivo (.xlsx, .xls ou .csv).",
                },
                "sheet_name": {
                    "type": "string",
                    "description": (
                        "Nome da aba a ler. Se omitido, usa a primeira aba "
                        "(ignorado para CSV)."
                    ),
                },
                "n_rows": {
                    "type": "integer",
                    "description": "Quantas linhas mostrar na amostra (padrão: 10).",
                },
            },
            "required": ["path"],
        },
    },
    {
        "type": "function",
        "name": "read_sheet",
        "description": (
            "Lê os dados de uma aba (ou CSV) de forma paginada. Use start_row "
            "e max_rows para percorrer planilhas grandes aos poucos, e columns "
            "para trazer só as colunas relevantes."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Caminho do arquivo (.xlsx, .xls ou .csv).",
                },
                "sheet_name": {
                    "type": "string",
                    "description": (
                        "Nome da aba a ler. Se omitido, usa a primeira aba "
                        "(ignorado para CSV)."
                    ),
                },
                "start_row": {
                    "type": "integer",
                    "description": "Linha inicial (0-indexado). Padrão: 0.",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "Máximo de linhas a retornar nesta chamada. Padrão: 200.",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Lista de colunas a retornar. Se omitido, retorna todas.",
                },
            },
            "required": ["path"],
        },
    },
    {
        "type": "function",
        "name": "search_in_sheet",
        "description": (
            "Busca um valor (texto ou número) dentro de uma aba/CSV, em uma "
            "coluna específica ou em todas. Útil para achar uma linha "
            "específica sem precisar ler a planilha inteira."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Caminho do arquivo (.xlsx, .xls ou .csv).",
                },
                "query": {
                    "type": "string",
                    "description": "Valor a buscar (substring, sem diferenciar maiúsculas/minúsculas).",
                },
                "sheet_name": {
                    "type": "string",
                    "description": (
                        "Nome da aba onde buscar. Se omitido, usa a primeira aba "
                        "(ignorado para CSV)."
                    ),
                },
                "column": {
                    "type": "string",
                    "description": "Coluna onde buscar. Se omitido, busca em todas as colunas.",
                },
                "max_matches": {
                    "type": "integer",
                    "description": "Máximo de linhas encontradas a retornar. Padrão: 50.",
                },
            },
            "required": ["path", "query"],
        },
    },
]
 